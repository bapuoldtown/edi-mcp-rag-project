import logging
from pathlib import Path 
from typing import List
import polars as pl 
from .base_parser import BaseParser 
from .models import ParsedDocument, ParserConfig, Table, DocumentType, TableCell

logger = logging.getLogger(__name__)


class ExcelParser(BaseParser):
    SUPPORTED_EXTENSIONS = {'xlsx', 'xls', 'csv', 'tsv'}
    
    def can_parse(self, file_path: str) -> bool:
        ext = self.get_file_extension(file_path)
        return ext in self.SUPPORTED_EXTENSIONS
    
    def parse(self, file_path: str) -> ParsedDocument:
        if not self.validate_file(file_path=file_path):
            raise ValueError(f"Invalid file: {file_path}")
        ext = self.get_file_extension(file_path=file_path)
        logger.info(f"Parsing {ext.upper()} file: {file_path}")
        
        try:
            if ext == 'csv':
                return self._parse_csv(file_path)
            elif ext == 'tsv':
                return self._parse_tsv(file_path)
            else:
                return self._parse_excel(file_path)
        except Exception as e:
            logger.error(f"Error parsing file: {e}")
            raise
                
    def _parse_csv(self, file_path: str) -> ParsedDocument:
        """Parse CSV file using Polars"""
        try:
            # Read CSV with Polars (super fast!)
            df = pl.read_csv(file_path, ignore_errors=True)
            return self._create_document_from_dataframe(file_path, df, DocumentType.CSV)
        except Exception as e:
            logger.error(f"CSV parsing error: {e}")
            # Fallback: try reading as text
            return self._parse_as_text(file_path, DocumentType.CSV)
    
    def _parse_tsv(self, file_path: str) -> ParsedDocument:
        """Parse TSV file using Polars"""
        try:
            df = pl.read_csv(file_path, separator='\t', ignore_errors=True)
            return self._create_document_from_dataframe(file_path, df, DocumentType.CSV)
        except Exception as e:
            logger.error(f"TSV parsing error: {e}")
            return self._parse_as_text(file_path, DocumentType.CSV)
    
    def _parse_excel(self, file_path: str) -> ParsedDocument:
        """Parse Excel file using Polars"""
        try:
            # Polars can read Excel files!
            # If multiple sheets, we'll read the first one by default
            df = pl.read_excel(file_path, sheet_id=0)
            
            # Try to get all sheet names for metadata
            metadata = {
                'sheet_name': 'Sheet1',  # Default
                'format': 'excel'
            }
            
            return self._create_document_from_dataframe(
                file_path, 
                df, 
                DocumentType.EXCEL,
                metadata
            )
        except Exception as e:
            logger.error(f"Excel parsing error: {e}")
            # Try with openpyxl as fallback
            return self._parse_excel_fallback(file_path)
    
    def _parse_excel_fallback(self, file_path: str) -> ParsedDocument:
        """Fallback Excel parser using openpyxl"""
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(file_path, read_only=True, data_only=True)
            sheet = wb.active
            
            # Extract data
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append([str(cell) if cell is not None else "" for cell in row])
            
            if not data:
                return self.create_parsed_document(
                    file_path,
                    DocumentType.EXCEL,
                    text="Empty spreadsheet"
                )
            
            # Convert to Polars DataFrame
            headers = data[0]
            rows = data[1:]
            
            df = pl.DataFrame(
                {header: [row[i] if i < len(row) else "" for row in rows] 
                 for i, header in enumerate(headers)}
            )
            
            return self._create_document_from_dataframe(file_path, df, DocumentType.EXCEL)
            
        except Exception as e:
            logger.error(f"Excel fallback parsing error: {e}")
            return self.create_parsed_document(
                file_path,
                DocumentType.EXCEL,
                text=f"Error parsing Excel file: {e}"
            )
    
    def _create_document_from_dataframe(
        self, 
        file_path: str, 
        df: pl.DataFrame,
        doc_type: DocumentType,
        metadata: dict = None
    ) -> ParsedDocument:
        """Convert Polars DataFrame to ParsedDocument"""
        
        # Extract table
        table = self._dataframe_to_table(df)
        
        # Create text representation
        text = self._dataframe_to_text(df)
        
        # Metadata
        meta = metadata or {}
        meta.update({
            'rows': df.height,
            'columns': df.width,
            'column_names': df.columns,
            'shape': f"{df.height} x {df.width}"
        })
        
        return self.create_parsed_document(
            file_path=file_path,
            document_type=doc_type,
            text=text,
            tables=[table],
            metadata=meta
        )
    
    def _dataframe_to_table(self, df: pl.DataFrame) -> Table:
        """Convert Polars DataFrame to Table object"""
        headers = df.columns
        
        # Convert to list of lists (rows)
        rows = []
        for row in df.iter_rows():
            rows.append([str(cell) for cell in row])
        
        return Table(
            headers=headers,
            rows=rows,
            num_rows=df.height,
            num_columns=df.width
        )
    
    def _dataframe_to_text(self, df: pl.DataFrame) -> str:
        """Convert DataFrame to readable text"""
        # Use Polars' built-in string representation
        return str(df)
    
    def _parse_as_text(self, file_path: str, doc_type: DocumentType) -> ParsedDocument:
        """Fallback: parse as plain text"""
        try:
            with open(file_path, 'r', encoding=self.config.encoding) as f:
                text = f.read()
            return self.create_parsed_document(file_path, doc_type, text=text)
        except Exception as e:
            logger.error(f"Text fallback error: {e}")
            return self.create_parsed_document(
                file_path,
                doc_type,
                text=f"Error reading file: {e}"
            )